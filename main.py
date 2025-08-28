#region imports
from re import S
import shutil
import pyodbc
import json
from openpyxl import load_workbook
from typing import Any
import os
#endregion


#region functions

def read_excel(excel_path: str) -> list[Any]:
    """
    Reads an Excel file and returns a list of worksheet objects.

    Args:
        excel_path (str): The file path to the Excel workbook.

    Returns:
        list[Any]: A list containing worksheet objects from the workbook. 
                   If the file path contains '~$', an empty list is returned.

    Note:
        The function skips files that appear to be temporary Excel files (containing '~$').
    """
    data:list[Any]= [] 
    if "~$" not in excel_path:
        wb = load_workbook(excel_path)
        sheets = wb.sheetnames
        for sheet in sheets:
            data.append(wb[sheet]) 
    return data

#endregion
#region Class
class sql_data_handler():
    #region setup
    def __init__(self, config: str, datas: list[list[Any]], paths: list[str]) -> None:
        self.config_path = config
        self.paths = paths
        self.config: dict[str,dict[str,str]] = {}
        self.dbconfig: dict[str,str] = {}
        self.col_names: list[str] = []
        self.cursor: pyodbc.Cursor
        self.server = ""
        self.database = ""
        self.username = ""
        self.password = ""
        self.driver = "{ODBC Driver 18 for SQL Server}"
        self.trust_server_certificate = "yes"
        self.encrypt = "yes"
        self.timeout = 30
        
        
        self.chem_data: list[list[str]] = []
        self.chemistry_A_data: list[str] = []
        self.chemistry_B_data: list[str] = []

        self.alphabet = [chr(i) for i in range(ord('A'), ord('Z')+1)]

        self.excel_datas = datas
        
        self.read_config()
    
    def read_config(self) -> None:
        with open(self.config_path, 'r') as f:
            self.config = json.load(f)
        self.dbconfig = self.config.get("Database_Config", {})
        self.server = self.dbconfig.get("host","localhost")
        self.database = self.dbconfig.get("db","cvd_test")
        self.driver = self.dbconfig.get("driver","{ODBC Driver 17 for SQL Server}")
        self.username = self.dbconfig.get("username", "")
        if self.username == "": self.username = input("No username in config please enter:")
        self.password = self.dbconfig.get("password", "")
        if self.password == "": self.password = input("No password in config please enter:")
    #endregion
    
    #reguib sql and excel utils
  
    def section_to_cols(self,  data: Any, alpha_range:list[str], number_range:list[list[int]], header:str, super_header_loc:list[str] = []) -> list[str]:
        complex_loc_locations = self.complex_addy(alpha_range, number_range, header)
        col_names = self.gen_col_names(data, complex_loc_locations, super_header_loc)
        return col_names

    def gen_col_names(self, data: Any, complex_col_name_locations:list[tuple[str,list[str]]], super_header_loc:list[str] = []) -> list[str]:
        cols:list[str] = []
        for complex_col in complex_col_name_locations:
            if super_header_loc != []:
                super_header = [data[loc].value for loc in super_header_loc]
                prefix = (".").join(super_header)
                
                prefix:str = prefix + "." + str(data[complex_col[0]].value)
            else:
                prefix = str(data[complex_col[0]].value)
            for col in complex_col[1]:
                if "Comments:" in str(data[col].value):
                    data_str =str(data[col].value)
                    data_str = data_str[:data_str.index(":")]
                    name = f"{prefix}.{data_str}"
                else:
                    name = f"{prefix}.{data[col].value}"
                name = name.replace(":","")
                name = name.replace("©", "C")
                name = name.replace(" ","")
                name = name.replace("Δ", "Delta")
                name = name.strip()
                if ".None" not in name:
                    self.col_names.append(name)
                    cols.append(name)
        return cols
    
    def complex_addy(self, alpha_range:list[str], number_range:list[list[int]], header:str) -> list[tuple[str,list[str]]]:
        i = 0 
        addys: list[str] = []
        for letter in alpha_range:
            for number in range(number_range[i][0], number_range[i][1] + 1):
                addy = letter + str(number)
                addys.append(addy)
            i += 1

        complex_col_name_locations: list[tuple[str,list[str]]] = [(header, addys)]
        

        return complex_col_name_locations
    
    def table_query_builder(self, table_name:str, cols:list[str], data_types:list[str]) -> None:
        """
        Builds a SQL CREATE TABLE query string for the specified table name, columns, and data types.
        Args:
            table_name (str): The name of the table to be created.
            cols (list[str]): A list of column names for the table.
            data_types (list[str]): A list of data types corresponding to each column.
        Returns:
            str: A SQL query string for creating the table with the specified columns and data types.
        Note:
            Only columns whose data type string contains both '(' and ')' will be included in the query.
        """

        query_pre:str = f"CREATE TABLE \"{table_name}\" ("
        temp:str = ""
        query_as_list:list[str] = []
        i = 0 
        for col in cols:
            if "(" in data_types[i] and ")" in data_types[i]:temp = f"\"{col}\" {data_types[i]}"
            query_as_list.append(temp)
            i += 1
            
        query_col = ",".join(query_as_list)
        query = f"{query_pre} {query_col})"

        return query

    def pull_chem_data(self, chem_col: list[str], chem_data: list[str|None]) -> list[list[str]]:
        cols:list[str] = []
        data:list[list[str]] = []

        i = 0
        for chem in chem_col:
            cols.append(chem)
            if "TotalVaporizerCarrierGasFlow" not in chem:
                temp:list[str] = []
                for point in chem_data[i * 5:(i+1) * 5]:
                    if point != None: temp.append(str(point))
                    else: temp.append("No Value")
                data.append(temp)
            else:
                if chem_data[-1] != None: data.append([chem_data[-1]])
                else: data.append([""])
            i += 1
        
        return data
    
    def write(self):
        chem_A_Data = self.pull_chem_data(self.chemistry_A, self.chem_data[0])
        chem_B_Data = self.pull_chem_data(self.chemistry_B, self.chem_data[1])
        query = f"INSERT INTO \"{self.table_name}\"("
        end = "("
        query_list = [f"\"{x}\"" for x in self.col_names if x not in self.chemistry_A and x not in self.chemistry_B]
        for name in self.col_names_basic:
            query_list.append(f"\"{name}\"")
        for chem in self.chemistry_A:
            query_list.append(f"\"{chem}\"")
        for chem in self.chemistry_B:
            query_list.append(f"\"{chem}\"")
        query_str = (',').join(query_list)
        query += query_str + ") "
        end_list = [f"\'{x}\'" for x in self.data_out]
        for point in self.data_basic:
            end_list.append(f"\'{point}\'")
        for chem in chem_A_Data:
            chem_str = (',').join(chem)
            end_list.append(f"\'{chem_str}\'")
        for chem in chem_B_Data:
            chem_str = (',').join(chem)
            end_list.append(f"\'{chem_str}\'")
        end = (',').join(end_list)
        end = "(" + end + ")"
        query += " VALUES " + end 

        self.cursor.execute(query)
        self.sql.commit()

    #endregion
    
    #region file management
    def move_file(self,cur_dir: str):
        done_dir = "processed"
        sheet_path = os.path.basename(cur_dir)
        dest_dir = os.path.join(done_dir, sheet_path)
        print(dest_dir)
        shutil.move(cur_dir, dest_dir)
    #endregion
    #region excel to sql
    
    def gen_all_cols(self, data:Any) -> None:
        # for data in self.excel_datas:
        self.col_names: list[str] = []
        #building complex column names
            
        #complex data
        # pre_coating_check A7
        precoat_alpha_range: list[str] = ["A", "C", "E", "G"]#[self.alphabet[i] for i in range(self.alphabet.index("A"), self.alphabet.index("G")+1)]
        precoat_number_range:list[list[int]] = [[9,14]] * (len(precoat_alpha_range) - 1)
        precoat_number_range.append([9,11])
        self.pre_coat = self.section_to_cols(data, precoat_alpha_range, precoat_number_range, "A7", [])
        #Conveyor Furnace Setpoints L7
        conveyor_alpha_range: list[str] = ["L"]
        conveyor_number_range:list[list[int]] = [[9,19]]
        self.conveyor = self.section_to_cols(data, conveyor_alpha_range, conveyor_number_range, "L7", [])
        #conveyor furnace settings
        cfs_super = ["A17"]
        #recipe A18
        cfs_alpha_range: list[str]= ["A"]
        cfs_number_range: list[list[int]]= [[19,25]]
        self.cfs_rec = self.section_to_cols(data, cfs_alpha_range, cfs_number_range, "A18", cfs_super)
        #N2 flows F18
        cfs_alpha_range = ["E"]
        cfs_number_range = [[19,24]]
        self.cfs_Ntwo = self.section_to_cols(data, cfs_alpha_range, cfs_number_range, "F18", cfs_super)
        #cart A
        cartA_super_headers = ["A28"]
        #exhaustBlower A30
        eb_alpharange = ["C","D","E","F","G","H"]
        eb_number_range = [[29,29]] * len(eb_alpharange)
        self.exhaust_blower_A = self.section_to_cols(data, eb_alpharange, eb_number_range, "A30", cartA_super_headers)
        #condenser temp L29
        ct_alpha_range = ["L"]
        ct_number_range = [[30,31]]
        self.condenser_temp_A = self.section_to_cols(data, ct_alpha_range, ct_number_range, "L29", cartA_super_headers)
        #exhaust_flow  A31
        af_alpharange = ["C","D","E","F","G","H"]
        af_number_range =[[29,29]] * len(af_alpharange)
        self.exhaust_flow_A = self.section_to_cols(data, af_alpharange, af_number_range, "A31", cartA_super_headers)
        #coater Inlet Mixer B33
        cim_alpha_range = ["C","D"]
        cim_number_range = [[33,33]] * len(cim_alpha_range)
        self.cim_A = self.section_to_cols(data, cim_alpha_range, cim_number_range, "B33", cartA_super_headers)
        #Coater Inlet Line F33
        cil_alpha_range = ["G","H"]
        cil_number_range = [[33,33]] * len(cil_alpha_range)
        self.cil_A = self.section_to_cols(data, cil_alpha_range, cil_number_range, "F33", cartA_super_headers)
        #bypass temp J33
        bt_alpha_range = ["K","L"]
        bt_number_range = [[33,33]] * len(bt_alpha_range)
        self.bypass_temp_A = self.section_to_cols(data, bt_alpha_range, bt_number_range, "J33", cartA_super_headers)
        #chemistry A37
        chem_alpha_range = ["A","B","C","D","E","F","G","I","J","K","L","M","F"]
        chem_number_range = [[38,38]] * (len(chem_alpha_range) -1)
        chem_number_range.append([44,44])
        self.chemistry_A = self.section_to_cols(data, chem_alpha_range, chem_number_range, "A37", cartA_super_headers)
        #coater Temperature P29
        ct_alpha_range = ["P","P"]
        ct_number_range = [[30,34],[36,36]]
        self.coater_temp_A = self.section_to_cols(data, ct_alpha_range, ct_number_range, "P29", cartA_super_headers)
        #TFE
        cartA_super_headers.append("P38")
        #tfe oil jacket Q39
        tfe_alpha_range = ["P"]
        tfe_number_range = [[40,41]]
        self.tfe_oil_jacket = self.section_to_cols(data, tfe_alpha_range, tfe_number_range, "Q39", cartA_super_headers)
        #tfe line 1 R39
        tfe_alpha_range = ["P"]
        tfe_number_range = [[40,41]]
        self.tfe_line_one = self.section_to_cols(data, tfe_alpha_range, tfe_number_range, "R39", cartA_super_headers)
        #tfe line 2 S39
        tfe_alpha_range = ["P"]
        tfe_number_range = [[40,41]]
        self.tfe_line_two = self.section_to_cols(data, tfe_alpha_range, tfe_number_range, "S39", cartA_super_headers)
        #init comments A47
        ic_alpha_range = ["A"]    
        ic_number_range = [[47,47]]
        self.ic_comments_A = self.section_to_cols(data, ic_alpha_range, ic_number_range, cartA_super_headers[0], [])
        #final_comments F47
        fc_alpha_range = ["F"]
        fc_number_range = [[47,47]]
        self.fc_comments_A = self.section_to_cols(data, fc_alpha_range, fc_number_range, cartA_super_headers[0], [])

        #cart B
        cartB_super_headers = ["U28"]
        #exhaustBlower U30
        eb_alpharange = ["W","X","Y","Z","AA","AB"]
        eb_number_range = [[29,29]] * len(eb_alpharange)
        self.exhaust_blower_B = self.section_to_cols(data, eb_alpharange, eb_number_range, "U30", cartB_super_headers)
        #condenser temp L29
        ct_alpha_range = ["L"]
        ct_number_range = [[30,31]]
        self.condenser_temp_B = self.section_to_cols(data, ct_alpha_range, ct_number_range, "L29", cartB_super_headers)
        #exhaust_flow  A31
        af_alpharange = ["C","D","E","F","G","H"]
        af_number_range =[[29,29]] * len(af_alpharange)
        self.exhaust_flow_B = self.section_to_cols(data, af_alpharange, af_number_range, "A31", cartB_super_headers)
        #coater Inlet Mixer B33
        cim_alpha_range = ["C","D"]
        cim_number_range = [[33,33]] * len(cim_alpha_range)
        self.cim_B = self.section_to_cols(data, cim_alpha_range, cim_number_range, "B33", cartB_super_headers)
        #Coater Inlet Line F33
        cil_alpha_range = ["G","H"]
        cil_number_range = [[33,33]] * len(cil_alpha_range)
        self.cil_B = self.section_to_cols(data, cil_alpha_range, cil_number_range, "F33", cartB_super_headers)
        #bypass temp J33
        bt_alpha_range = ["K","L"]
        bt_number_range = [[33,33]] * len(bt_alpha_range)
        self.bypass_temp_B = self.section_to_cols(data, bt_alpha_range, bt_number_range, "J33", cartB_super_headers)
        #chemistry A37
        chem_alpha_range = ["A","B","C","D","E","F","G","I","J","K","L","M","F"]
        chem_number_range = [[38,38]] * (len(chem_alpha_range) -1)
        chem_number_range.append([44,44])
        self.chemistry_B = self.section_to_cols(data, chem_alpha_range, chem_number_range, "A37", cartB_super_headers)
        #coater Temperature P29
        ct_alpha_range = ["P","P"]
        ct_number_range = [[30,34],[36,36]]
        self.coater_temp_B = self.section_to_cols(data, ct_alpha_range, ct_number_range, "P29", cartB_super_headers)
        #init comments A47
        ic_alpha_range = ["A"]    
        ic_number_range = [[47,47]]
        self.ic_comments_B = self.section_to_cols(data, ic_alpha_range, ic_number_range, cartB_super_headers[0], [])
        #final_comments F47
        fc_alpha_range = ["F"]
        fc_number_range = [[47,47]]
        self.fc_comments_B = self.section_to_cols(data, fc_alpha_range, fc_number_range, cartB_super_headers[0], [])
        
    def build_table(self, data:Any):
        # for data in self.excel_datas:
        self.col_names_basic: list[str] = []
        sample:Any = data["I3"].value #type:ignore
        run:Any = data["N3"].value #type:ignore
        self.table_name:str = f"cvd_{sample}_{run}"
        data_holder = self.cursor.execute("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE'")
        self.tables = [x[2] for x in data_holder]
        col_name_basic_addy:list[str] = ["A3", "H3", "M3", "A5", "H5"]
        col_data_basic_addy:list[str] = ["B2", "I3", "N3", "B5", "I5"]
        self.col_names_basic = [data[loc].value for loc in col_name_basic_addy]
        self.data_basic = [data[loc].value for loc in col_data_basic_addy]
        i = 0
        for col in self.col_names_basic:
            if col[-1] == " ": 
                col = col[:-1]
            col = col.replace(" ","_")
            col = col.replace(":","")
            self.col_names_basic[i] = col
            i += 1
        col_data_types:list[str] = ["VARCHAR(255)", "VARCHAR(255)", "VARCHAR(255)", "VARCHAR(255)", "VARCHAR(255)"]

        table_query =self.table_query_builder(self.table_name, self.col_names_basic, col_data_types) 

        if self.table_name not in self.tables:
            self.cursor.execute(table_query)
            self.cursor.commit()
            print(f"Table {self.table_name} created")
    
    def build_cols(self):
        query = f'SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = \'{self.table_name}\''
        
        self.cursor.execute(query)
        results = [x[0] for x in self.cursor.fetchall()]
        print(self.table_name)
        
        query = f"ALTER TABLE \"{self.table_name}\" ADD "
        col_data_types:list[str] = ["VARCHAR(255)"] * (len(self.col_names) -2)
        col_data_types.append("VARCHAR(MAX)")
        col_data_types.append("VARCHAR(MAX)")
        query_list: list[str] = []
        i = 0
        for col in self.col_names:
            if col not in results:
                query_list.append(f'"{col}" {col_data_types[i]}')
            i += 1
        if query_list != []:
            query_str = (",").join(query_list)
            query += query_str
            self.cursor.execute(query)
        self.cursor.commit()
    
    def execute(self):
        i = 0
        for data in self.excel_datas: 
            self.chem_data = []
            for sheet in data:
                self.gen_all_cols(sheet)
                self.build_table(sheet)
                self.build_cols()
                self.gen_all_data_addy(sheet)
                self.write()
                
            self.move_file(self.paths[i])
            i += 1  
        
    #endregion
    
    #region data processing
    def get_data(self, addys: list[str], data:Any, chem:bool = False) -> list[str]:
        local_out: list[str] = []
        for addy in addys:
            d = (data[addy].value)
            local_out.append(d)
            if not chem:
                self.data_out.append(d)
        if chem:
            # if len(self.chemistry_A_data) != 0:
            self.chem_data.append(local_out)
        return local_out
    
    def get_comments(self, comments:str|None) -> str:
        if comments != None:
            temp = [x for x in comments.split("\n") if x != ""]
            out = [x for x in temp if "Initial Comments" not in x and "Final Comments" not in x]
            comment = ("; ").join(out)
            self.data_out.append(comment)
            return comment
        else:
            self.data_out.append("")
            return ""

    def gen_all_data_addy(self, data:Any) -> None:
        # for data in self.excel_datas:
        self.data_out: list[str] = []
        
        #building complex column names
            
        #complex data
        # pre_coating_check A7
        precoat_alpha_range: list[str] = ["B", "D", "F", "J"]#[self.alphabet[i] for i in range(self.alphabet.index("A"), self.alphabet.index("G")+1)]
        precoat_number_range:list[list[int]] = [[9,14]] * (len(precoat_alpha_range) - 1)
        precoat_number_range.append([9,11])
        self.pre_coat_data_addy = self.complex_addy(precoat_alpha_range, precoat_number_range, "A7")
        self.pre_coat_data = self.get_data(self.pre_coat_data_addy[0][1], data)
        #Conveyor Furnace Setpoints L7
        conveyor_alpha_range: list[str] = ["M"]
        conveyor_number_range:list[list[int]] = [[9,19]]
        self.conveyor_data_addy = self.complex_addy(conveyor_alpha_range, conveyor_number_range, "L7")
        self.conveyor_data = self.get_data(self.conveyor_data_addy[0][1], data)
        #conveyor furnace settings
        cfs_super = ["A17"]
        #recipe A18
        cfs_alpha_range: list[str]= ["B"]
        cfs_number_range: list[list[int]]= [[19,25]]
        self.cfs_data_addy = self.complex_addy(cfs_alpha_range, cfs_number_range, "A18")
        self.cfs_rec_data = self.get_data(self.cfs_data_addy[0][1], data)
        #N2 flows F18
        cfs_alpha_range = ["G"]
        cfs_number_range = [[19,24]]
        self.cfs_Ntwo_data_addy = self.complex_addy(cfs_alpha_range, cfs_number_range, "F18")
        self.cfs_Ntwo_data = self.get_data(self.cfs_Ntwo_data_addy[0][1], data)
        #cart A
        cartA_super_headers = ["A28"]
        #exhaustBlower A30
        eb_alpharange = ["C","D","E","F","G","H"]
        eb_number_range = [[30,30]] * len(eb_alpharange)
        self.exhaust_blower_A_data_addy = self.complex_addy(eb_alpharange, eb_number_range, "A30")
        self.exhaust_blower_A_data = self.get_data(self.exhaust_blower_A_data_addy[0][1], data)
        cut = len(self.exhaust_blower_A) - len(self.exhaust_blower_A_data)
        if cut < 0: 
            self.data_out = self.data_out[:cut]
        #condenser temp L29
        ct_alpha_range = ["M"]
        ct_number_range = [[30,31]]
        self.condenser_temp_A_data_addy = self.complex_addy(ct_alpha_range, ct_number_range, "L29")
        self.condenser_temp_A_data = self.get_data(self.condenser_temp_A_data_addy[0][1], data)
        #exhaust_flow  A31
        af_alpharange = ["C","D","E","F","G","H"]
        af_number_range =[[31,31]] * len(af_alpharange)
        self.exhaust_flow_A_data_addy = self.complex_addy(af_alpharange, af_number_range, "A31")
        self.exhaust_flow_A_data = self.get_data(self.exhaust_flow_A_data_addy[0][1], data)
        cut = len(self.exhaust_flow_A) - len(self.exhaust_flow_A_data)
        if cut < 0: 
            self.data_out = self.data_out[:cut]
        #coater Inlet Mixer B33
        cim_alpha_range = ["C","D"]
        cim_number_range = [[34,34]] * len(cim_alpha_range)
        self.cim_A_data_addy = self.complex_addy(cim_alpha_range, cim_number_range, "B33")
        self.cim_A_data = self.get_data(self.cim_A_data_addy[0][1], data)
        #Coater Inlet Line F33
        cil_alpha_range = ["G","H"]
        cil_number_range = [[34,34]] * len(cil_alpha_range)
        self.cil_A_data_addy = self.complex_addy(cil_alpha_range, cil_number_range, "F33")
        self.cil_A_data = self.get_data(self.cil_A_data_addy[0][1], data)
        #bypass temp J33
        bt_alpha_range = ["K","L"]
        bt_number_range = [[34,34]] * len(bt_alpha_range)
        self.bypass_temp_A_data_addy = self.complex_addy(bt_alpha_range, bt_number_range, "J33")
        self.bypass_temp_A_data = self.get_data(self.bypass_temp_A_data_addy[0][1], data)
        #chemistry A37
        chem_alpha_range = ["A","B","C","D","E","F","G","I","J","K","L","M","H"]
        chem_number_range = [[39,43]] * (len(chem_alpha_range) -1)
        chem_number_range.append([44,44])
        self.chemistry_A_data_addy = self.complex_addy(chem_alpha_range, chem_number_range, "A37")
        self.chemistry_A_data = self.get_data(self.chemistry_A_data_addy[0][1], data, True) 
        #coater Temperature P29
        ct_alpha_range = ["Q","Q"]
        ct_number_range = [[30,34],[36,36]]
        self.coater_temp_A_data_addy = self.complex_addy(ct_alpha_range, ct_number_range, "P29")
        self.coater_temp_A_data = self.get_data(self.coater_temp_A_data_addy[0][1], data)
        #TFE
        cartA_super_headers.append("P38")
        #tfe oil jacket Q39
        tfe_alpha_range = ["Q"]
        tfe_number_range = [[40,41]]
        self.tfe_oil_jacket_data_addy = self.complex_addy(tfe_alpha_range, tfe_number_range, "Q39")
        self.tfe_oil_jacket_data = self.get_data(self.tfe_oil_jacket_data_addy[0][1], data)
        #tfe line 1 R39
        tfe_alpha_range = ["R"]
        tfe_number_range = [[40,41]]
        self.tfe_line_one_data_addy = self.complex_addy(tfe_alpha_range, tfe_number_range, "R39")
        self.tfe_line_one_data = self.get_data(self.tfe_line_one_data_addy[0][1], data)
        #tfe line 2 S39
        tfe_alpha_range = ["S"]
        tfe_number_range = [[40,41]]
        self.tfe_line_two_data_addy = self.complex_addy(tfe_alpha_range, tfe_number_range, "S39")
        self.tfe_line_two_data = self.get_data(self.tfe_line_two_data_addy[0][1], data)
        #init comments A47
        ic_alpha_range = ["A"]    
        ic_number_range = [[48,48]]
        self.ic_comments_A_data = self.get_comments(data[self.complex_addy(ic_alpha_range, ic_number_range, "A47")[0][1][0]].value)
        
        # self.ic_comments_A = self.section_to_cols(data, ic_alpha_range, ic_number_range, cartA_super_headers[0], [])
        #final_comments F47
        fc_alpha_range = ["F"]
        fc_number_range = [[48,48]]
        self.fc_comments_A_data = self.get_comments(data[self.complex_addy(fc_alpha_range, fc_number_range, "F47")[0][1][0]].value)
        # self.fc_comments_A = self.section_to_cols(data, fc_alpha_range, fc_number_range, cartA_super_headers[0], [])

        #cart B
        #exhaustBlower U30
        eb_alpharange = ["W","X","Y","Z","AA","AB"]
        eb_number_range = [[30,30]] * len(eb_alpharange)
        self.exhaust_blower_B_data = self.get_data(self.complex_addy(eb_alpharange, eb_number_range, "U30")[0][1], data)
        cut = len(self.exhaust_blower_B) - len(self.exhaust_blower_B_data)
        if cut < 0: 
            self.data_out = self.data_out[:cut]
        #condenser temp L29
        ct_alpha_range = ["AG"]
        ct_number_range = [[30,31]]
        self.condenser_temp_B_data = self.get_data(self.complex_addy(ct_alpha_range, ct_number_range, "AG30")[0][1], data)
        #exhaust_flow  U31
        af_alpharange = ["W","X","Y","Z","AA","AB"]
        af_number_range =[[30,30]] * len(af_alpharange)
        self.exhaust_flow_B_data = self.get_data(self.complex_addy(af_alpharange, af_number_range, "U31")[0][1], data)
        cut = len(self.exhaust_flow_B) - len(self.exhaust_flow_B_data)
        if cut < 0:
            self.data_out = self.data_out[:cut]
        #coater Inlet Mixer V33
        cim_alpha_range = ["W","X"]
        cim_number_range = [[34,34]] * len(cim_alpha_range)
        self.coater_inlet_mixer_B_data = self.get_data(self.complex_addy(cim_alpha_range, cim_number_range, "V33")[0][1], data)
        #Coater Inlet Line Z33
        cil_alpha_range = ["AA","AB"]
        cil_number_range = [[34,34]] * len(cil_alpha_range)
        self.cil_B_data = self.get_data(self.complex_addy(cil_alpha_range, cil_number_range, "Z33")[0][1], data)
        #bypass temp J33
        bt_alpha_range = ["AE","AF"]
        bt_number_range = [[34,34]] * len(bt_alpha_range)
        self.bypass_temp_B_data = self.get_data(self.complex_addy(bt_alpha_range, bt_number_range, "AE33")[0][1], data)
        #chemistry A37
        chem_alpha_range = ["U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG"]
        chem_number_range = [[39,43]] * (len(chem_alpha_range) -1)
        chem_number_range.append([45,45])
        self.chemistry_B_data = self.get_data(self.complex_addy(chem_alpha_range, chem_number_range, "U37")[0][1], data, True)
        #coater Temperature P29
        ct_alpha_range = ["AK","AK"]
        ct_number_range = [[30,34],[36,36]]
        self.coater_temp_B_data = self.get_data(self.complex_addy(ct_alpha_range, ct_number_range, "AK30")[0][1], data)
        #init comments A47
        ic_alpha_range = ["U"]    
        ic_number_range = [[48,48]]
        self.ic_comments_B_data = self.get_comments(data[self.complex_addy(ic_alpha_range, ic_number_range, "U47")[0][1][0]].value)
        #final_comments F47
        fc_alpha_range = ["Z"]
        fc_number_range = [[48,48]]
        print(data[self.complex_addy(fc_alpha_range, fc_number_range, "Z47")[0][1][0]].value)
        self.fc_comments_B_data = self.get_comments(data[self.complex_addy(fc_alpha_range, fc_number_range, "Z47")[0][1][0]].value)

    #endregion
    
    #region sql server connections  
    def connect(self) -> None:
        self.sql = pyodbc.connect(
            host=self.server,
            user=self.username,
            password=self.password,
            database=self.database,
            driver=self.driver,
            autocommit=True
        )
        self.cursor = self.sql.cursor()

    def close(self) -> None:
        if self.sql:
            self.sql.close()
    #endregion
    
#region entry point
if __name__ == "__main__":

    paths:list[str] = []
    datas:list[list[Any]] = []
    
    for _,_,files in os.walk("to_process"):
        for file in files:
            path = os.path.join("to_process",file)
            paths.append(path)
            if ".xlsx" in file: datas.append(read_excel(path))
    temp = sql_data_handler("config.json", datas, paths)

    temp.connect()
    temp.execute()
    temp.close()
    
    i = 0

    
#end region